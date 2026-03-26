import os
import re
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

class InvoiceParser:
    """发票文件解析器，实现文档2.1节功能"""
    
    # 匹配规则：dzfp_订单号_姓名_时间戳.pdf
    FILENAME_PATTERN = re.compile(r'^dzfp_([^_]+)_[^_]+_\d+\.pdf$')

    @staticmethod
    def parse_directory(dir_path):
        """
        遍历目录，解析PDF文件并生成字典
        :param dir_path: 发票存放目录
        :return: (dict {订单号: 绝对路径}, list [解析失败的文件名])
        """
        invoice_map = {}
        failed_files = []
        
        if not dir_path or not os.path.exists(dir_path):
            return {}, []

        for filename in os.listdir(dir_path):
            if not filename.lower().endswith('.pdf'):
                continue
            
            match = InvoiceParser.FILENAME_PATTERN.match(filename)
            if match:
                order_no = match.group(1)
                full_path = os.path.abspath(os.path.join(dir_path, filename))
                invoice_map[order_no] = full_path
            else:
                failed_files.append(filename)
                
        return invoice_map, failed_files

    @staticmethod
    def generate_excel(invoice_map, output_dir):
        """
        生成发票上传处理记录表（Excel）
        :param invoice_map: 解析出的发票字典
        :param output_dir: 指定的保存目录
        :return: 生成的Excel文件路径
        """
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        excel_name = f"发票上传处理记录表_{timestamp}.xlsx"
        
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        excel_path = os.path.join(output_dir, excel_name)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "处理记录"
        
        # 写入表头
        ws.append(["订单号", "文件路径", "处理状态"])
        
        # 写入初始数据
        for order_no, file_path in invoice_map.items():
            ws.append([order_no, file_path, "未处理"])
            
        wb.save(excel_path)
        return excel_path

    @staticmethod
    def update_excel_status(excel_path, order_no, status, message=""):
        if not excel_path or not os.path.exists(excel_path):
            return

        wb = load_workbook(excel_path)
        ws = wb.active

        header = [cell.value for cell in ws[1]]
        if "处理备注" not in header:
            ws.cell(row=1, column=4, value="处理备注")

        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=1).value) == str(order_no):
                ws.cell(row=row, column=3, value=status)
                ws.cell(row=row, column=4, value=message)
                break

        wb.save(excel_path)
